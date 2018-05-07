#encoding=utf-8

import re
import random
import copy
import openpyxl

shop_list = list()
locate_list = list()
type_list = list()
food_list = list()
mrt_list = list()
build_list = list()
training_data = open("./dataset/ijia_dataset_musk.iob", "r").readlines()
training_data_expend = open("ijia_dataset_musk_expend.iob", "w")

file = "entity_type.xlsx"

entity_type =openpyxl.load_workbook(file)
sheets = entity_type.sheetnames
sheet1 = sheets[1]
ws = entity_type.get_sheet_by_name(sheet1)


def add_sentance(list, label):
    print("========== 正在處理 " + label.encode("utf-8") + " ===========")
    for i in range(1,21):
        if ws[2][i].value == label:
            ws_colum = i
            #print(i)
            break
    else:
        print("no label match")

    label_utf8 = label.encode("utf-8")
    for row in range(3,22319):
        if row%500 == 0:
            print("已處理 " +str(row)+" 筆資料")

        word = ws[row][ws_colum].value
        if word:
            word = word.encode("utf-8")
            sentance_index = random.randint(0,len(list)-1)
            word_index = list[sentance_index][1].index(label_utf8)
            #expend_sentance = copy.deepcopy( list[sentance_index] )
            expend_sentance = list[sentance_index]
            expend_sentance[0][word_index] = word

            a = " ".join( expend_sentance[0] )
            b = " ".join( expend_sentance[1] )
            s = a + "\t" + b + " " + expend_sentance[2]

            training_data_expend.write( s )
        else:
            break



for t in training_data:
    #train_data = re.sub("\n", "", line).sub("\t", " ", line).split(" ")
    train_data = [t.split( "\t" )[0].split(" "), t.split("\t")[1].split(" ")[:-1], t.split("\t")[1].split(" ")[-1]]

    if train_data[2] == "找餐廳\n":
        if "B-餐廳需求::地點" in train_data[1]:
            locate_list.append(train_data)
        if "B-餐廳需求::地標" in train_data[1]:
            build_list.append(train_data)
        if "B-餐廳需求::捷運站" in train_data[1]:
            mrt_list.append(train_data)
        if "B-餐廳需求::店名" in train_data[1]:
            shop_list.append(train_data)
        if "B-餐廳需求::餐點" in train_data[1]:
            food_list.append(train_data)
        if "B-餐廳需求::類型" in train_data[1]:
            type_list.append(train_data)

add_sentance(shop_list,u"B-餐廳需求::店名")
add_sentance(type_list,u"B-餐廳需求::類型")
add_sentance(food_list,u"B-餐廳需求::餐點")
add_sentance(locate_list,u"B-餐廳需求::地點")
add_sentance(build_list,u"B-餐廳需求::地標")
add_sentance(mrt_list,u"B-餐廳需求::捷運站")